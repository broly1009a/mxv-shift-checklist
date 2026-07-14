import os
import sys
import json
import zipfile
import xml.etree.ElementTree as ET
import re
import openpyxl

# Reconfigure stdout to use utf-8 for Windows console/piping
sys.stdout.reconfigure(encoding='utf-8')

def parse_docx(docx_path, commodity_xlsx_path):
    if not os.path.exists(docx_path):
        return {"success": False, "error": f"Docx file not found: {docx_path}"}
    
    # 1. Load Commodity.xlsx margins to build lookup table
    current_margins = {}
    if os.path.exists(commodity_xlsx_path):
        try:
            wb = openpyxl.load_workbook(commodity_xlsx_path, read_only=True)
            sheet = wb.active
            
            headers = []
            for row in sheet.iter_rows(max_row=1, values_only=True):
                headers = [str(h).strip().lower() for h in row if h is not None]
                break
            
            # Find column indices
            code_col = -1
            margin_col = -1
            for idx, h in enumerate(headers):
                if "mã hàng hóa" in h or "symbol" in h:
                    code_col = idx
                elif "mức ký quỹ ban đầu" in h or "margin" in h:
                    margin_col = idx
            
            # If not found, use default indices
            if code_col == -1: code_col = 2
            if margin_col == -1: margin_col = 5
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > max(code_col, margin_col):
                    code = row[code_col]
                    margin = row[margin_col]
                    if code is not None and margin is not None:
                        code_str = str(code).strip().upper()
                        try:
                            # Normalize margin (remove commas)
                            margin_val = float(str(margin).replace(',', '').strip())
                            current_margins[code_str] = margin_val
                        except ValueError:
                            pass
        except Exception as e:
            # Non-blocking, print to stderr and continue
            print(f"Warning loading Commodity.xlsx: {e}", file=sys.stderr)

    # 2. Parse Docx File XML
    try:
        with zipfile.ZipFile(docx_path) as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            
            # 2.1 Extract Effective Date from text
            effective_date = None
            for paragraph in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                para_text = []
                for run in paragraph.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                    if run.text:
                        para_text.append(run.text)
                para_str = "".join(para_text)
                
                # Look for date pattern: "kể từ ngày 13/7/2026" or similar
                date_match = re.search(r'(?:hiệu lực|kể từ) ngày\s+([\d]{1,2}/[\d]{1,2}/[\d]{4})', para_str, re.IGNORECASE)
                if date_match:
                    effective_date = date_match.group(1)
                    break
            
            if not effective_date:
                # Try a broader regex
                for paragraph in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                    para_text = [r.text for r in paragraph.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t') if r.text]
                    para_str = "".join(para_text)
                    date_match = re.search(r'ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{4})', para_str, re.IGNORECASE)
                    if date_match:
                        effective_date = f"{date_match.group(1)}/{date_match.group(2)}/{date_match.group(3)}"
                        break
            
            session_str = f"Phiên T ngày {effective_date}" if effective_date else "Phiên T"

            # 2.2 Parse Tables
            tables = root.findall('.//w:tbl', ns)
            
            # Find the commodities table.
            # We look for a table containing headers like "Mức ký quỹ" and "Mã hàng hóa"
            commodities_tbl = None
            for tbl in tables:
                first_row = tbl.find('w:tr', ns)
                if first_row is not None:
                    cells = first_row.findall('w:tc', ns)
                    row_texts = []
                    for cell in cells:
                        cell_text = "".join([t.text for t in cell.findall('.//w:t', ns) if t.text])
                        row_texts.append(cell_text.lower())
                    
                    if any("mã hàng hóa" in txt for txt in row_texts) and any("ký quỹ" in txt for txt in row_texts):
                        commodities_tbl = tbl
                        break
            
            # If not found by header keyword, fallback to the 4th table (index 3)
            if commodities_tbl is None and len(tables) >= 4:
                commodities_tbl = tables[3]
            
            if commodities_tbl is None:
                return {"success": False, "error": "Commodities table not found in document"}
            
            rows = commodities_tbl.findall('w:tr', ns)
            if len(rows) < 2:
                return {"success": False, "error": "Commodities table is empty"}
            
            # Extract header columns mapping
            header_cells = rows[0].findall('w:tc', ns)
            header_texts = []
            for cell in header_cells:
                cell_text = "".join([t.text for t in cell.findall('.//w:t', ns) if t.text]).strip().lower()
                header_texts.append(cell_text)
            
            name_idx = -1
            code_idx = -1
            margin_idx = -1
            
            for idx, txt in enumerate(header_texts):
                if "tên hàng hóa" in txt or "tên sản phẩm" in txt:
                    name_idx = idx
                elif "mã hàng hóa" in txt or "mã sản phẩm" in txt:
                    code_idx = idx
                elif "mức ký quỹ" in txt or "ký quỹ ban đầu" in txt:
                    margin_idx = idx
            
            # Fallback indices
            if name_idx == -1: name_idx = 1
            if code_idx == -1: code_idx = 2
            if margin_idx == -1: margin_idx = 5
            
            filename = os.path.basename(docx_path)
            changes = []
            
            # Iterate through rows starting from index 1 (skip header)
            for r_idx in range(1, len(rows)):
                row = rows[r_idx]
                cells = row.findall('w:tc', ns)
                if len(cells) <= max(name_idx, code_idx, margin_idx):
                    continue
                
                # Check if this row has bold runs in key cells (Tên, Mã, Kí quỹ)
                is_bold = False
                for target_idx in [name_idx, code_idx, margin_idx]:
                    cell = cells[target_idx]
                    for p in cell.findall('w:p', ns):
                        for r in p.findall('w:r', ns):
                            rPr = r.find('w:rPr', ns)
                            if rPr is not None:
                                b = rPr.find('w:b', ns)
                                if b is not None:
                                    val = b.attrib.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                                    if val not in ('false', '0'):
                                        is_bold = True
                                        break
                        if is_bold:
                            break
                    if is_bold:
                        break
                
                if not is_bold:
                    continue
                
                # Extract text for Name, Code, New Margin
                cell_name_txt = "".join([t.text for t in cells[name_idx].findall('.//w:t', ns) if t.text]).strip()
                cell_code_txt = "".join([t.text for t in cells[code_idx].findall('.//w:t', ns) if t.text]).strip()
                cell_margin_txt = "".join([t.text for t in cells[margin_idx].findall('.//w:t', ns) if t.text]).strip()
                
                if not cell_code_txt:
                    continue
                
                # Parse new margin number
                try:
                    new_margin_val = float(cell_margin_txt.replace(',', '').strip())
                except ValueError:
                    continue
                
                code_normalized = cell_code_txt.strip().upper()
                old_margin_val = current_margins.get(code_normalized, new_margin_val) # Fallback to new if not found
                
                # If old margin equals new margin, it means the docx table has the same value as Commodity.xlsx,
                # but since it's bolded, we still treat it as a change (or they want to re-apply it).
                
                commodity_name_display = f"{cell_name_txt} ({cell_code_txt})" if cell_name_txt else cell_code_txt
                
                changes.append({
                    "commodity": commodity_name_display,
                    "oldMargin": old_margin_val,
                    "newMargin": new_margin_val,
                    "effectiveSession": session_str,
                    "comments": f"Trích xuất tự động từ Quyết định: {filename}"
                })
            
            return {
                "success": True,
                "filePath": docx_path,
                "effectiveSession": session_str,
                "changes": changes
            }
            
    except Exception as e:
        return {"success": False, "error": f"Failed to parse docx: {str(e)}"}

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(json.dumps({"success": False, "error": "Usage: python parse_margin_decision.py <docx_path> <commodity_xlsx_path>"}))
        sys.exit(1)
        
    docx_path = sys.argv[1]
    commodity_xlsx_path = sys.argv[2]
    
    result = parse_docx(docx_path, commodity_xlsx_path)
    print(json.dumps(result, ensure_ascii=False))
