#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_sheet_cloner.py
Tự động nhân bản (clone) và sinh Sheet tháng mới trên Ubuntu Linux / Windows
sử dụng openpyxl copy_worksheet mà không làm vỡ Shared Formulas hay định dạng XML.
Đã tối ưu chuẩn xác theo quy tắc cấu trúc file Thống kê Số Lot & Giá trị của MXV.
"""

import sys
import os
import re
import argparse
from datetime import datetime

# Thiết lập UTF-8 cho stdout/stderr để tránh lỗi charmap trên Windows console
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    try:
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import openpyxl

def parse_month_year_from_sheet_name(sheet_name: str):
    """
    Trích xuất (month, year) từ tên sheet dạng T08.2026 hoặc T08_2026
    """
    m = re.search(r'T(\d{1,2})[._](\d{4})', sheet_name, re.IGNORECASE)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def clone_month_sheet(excel_path: str, target_sheet_name: str, clean_data: bool = True) -> bool:
    if not os.path.exists(excel_path):
        print(f"[ERROR] File khong ton tai: {excel_path}", file=sys.stderr)
        return False

    try:
        print(f"[INFO] Dang mo file Excel: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        
        # 1. Kiểm tra nếu sheet đích đã tồn tại
        if target_sheet_name in wb.sheetnames:
            print(f"[INFO] Sheet '{target_sheet_name}' da ton tai san. Khong can tao moi.")
            return True

        # 2. Xác định tháng/năm mục tiêu từ tên sheet
        target_month, target_year = parse_month_year_from_sheet_name(target_sheet_name)
        if not target_month or not target_year:
            now = datetime.now()
            target_month = now.month
            target_year = now.year

        # 3. Tìm sheet nguồn (ưu tiên sheet tháng gần nhất)
        month_sheets = [s for s in wb.sheetnames if s.startswith('T') and ('.' in s or '_' in s)]
        
        if month_sheets:
            source_sheet_name = month_sheets[-1]
        else:
            source_sheet_name = wb.sheetnames[-1]

        source_sheet = wb[source_sheet_name]
        print(f"[INFO] Nhan ban cau truc tu Sheet mau: '{source_sheet_name}' -> '{target_sheet_name}'...")

        # 4. Sao chép nguyên vẹn 100% XML, mergeCells, Styles, Formulas
        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = target_sheet_name

        # 5. Cập nhật tiêu đề hiển thị tháng mới trong ô A1 (nếu có chuỗi tháng cũ)
        cell_a1 = new_sheet.cell(row=1, column=1)
        if cell_a1.value and isinstance(cell_a1.value, str):
            old_str = cell_a1.value
            new_str = re.sub(r'th[aá]ng\s+\d{1,2}/\d{4}', f'tháng {target_month:02d}/{target_year}', old_str, flags=re.IGNORECASE)
            new_sheet.cell(row=1, column=1).value = new_str

        # 6. Cập nhật mốc ngày đầu tháng (Ngày 1) cho công thức =WORKDAY(...)
        # - Trường hợp File Thống kê Số Lot: Ngày đầu tháng nằm ở ô B5
        # - Trường hợp File Thống kê Giá trị: Ngày đầu tháng nằm ở ô A6
        first_date_val = datetime(target_year, target_month, 1)

        cell_b5 = new_sheet.cell(row=5, column=2).value
        cell_a6 = new_sheet.cell(row=6, column=1).value

        if isinstance(cell_b5, datetime) or (isinstance(cell_b5, str) and re.match(r'^\d{4}-\d{2}-\d{2}', str(cell_b5))):
            new_sheet.cell(row=5, column=2).value = first_date_val
            print(f"[INFO] Cap nhat moc ngay dau thang tai B5: {first_date_val.strftime('%Y-%m-%d')}")

        if isinstance(cell_a6, datetime) or (isinstance(cell_a6, str) and re.match(r'^\d{4}-\d{2}-\d{2}', str(cell_a6))):
            new_sheet.cell(row=6, column=1).value = first_date_val
            print(f"[INFO] Cap nhat moc ngay dau thang tai A6: {first_date_val.strftime('%Y-%m-%d')}")

        # 7. Xóa trắng dữ liệu giao dịch cũ của các ngày trong tháng (giữ nguyên công thức và tiêu đề)
        if clean_data:
            print("[INFO] Dang don dep du lieu ngay cu (giu nguyen cong thuc)...")
            
            max_row = new_sheet.max_row
            max_col = new_sheet.max_column

            for r in range(5, max_row + 1):
                cell_a = new_sheet.cell(row=r, column=1).value
                cell_b = new_sheet.cell(row=r, column=2).value
                
                # Bỏ qua dòng TỔNG / TOTAL
                str_a = str(cell_a).upper() if cell_a else ""
                str_b = str(cell_b).upper() if cell_b else ""
                if "TỔNG" in str_a or "TOTAL" in str_a or "TỔNG" in str_b or "TOTAL" in str_b or "TONG" in str_a or "TONG" in str_b:
                    continue

                # Bỏ qua các dòng chỉ chứa nhãn/tiêu đề
                # Chỉ xóa các ô số liệu (cột 3 trở đi cho Lot, cột 2 trở đi cho Value)
                start_clean_col = 3 if cell_b5 is not None else 2
                for c in range(start_clean_col, max_col + 1):
                    cell = new_sheet.cell(row=r, column=c)
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Giữ nguyên toàn bộ ô công thức bắt đầu bằng '='
                        if not val_str.startswith('='):
                            cell.value = None

        # 8. Lưu lại file hoàn chỉnh
        print(f"[INFO] Dang luu file...")
        wb.save(excel_path)
        print(f"[SUCCESS] Da tu dong sinh Sheet moi '{target_sheet_name}' thanh cong 100%!")
        return True

    except Exception as e:
        print(f"[ERROR] Loi khi nhan ban Sheet: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return False

def main():
    parser = argparse.ArgumentParser(description="Tu dong nhan ban Sheet thang moi cho file Excel thong ke")
    parser.add_argument("--file", "-f", required=True, help="Duong dan toi file Excel")
    parser.add_argument("--sheet", "-s", required=True, help="Ten sheet moi can tao (vi du: T08.2026)")
    parser.add_argument("--no-clean", action="store_true", help="Khong xoa du lieu ngay cu")

    args = parser.parse_args()
    success = clone_month_sheet(args.file, args.sheet, clean_data=not args.no_clean)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
